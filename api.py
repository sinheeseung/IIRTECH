import urllib.parse
import urllib.request
import json
import openpyxl
import pandas as pd
import re

opendict_cert_key = 'FCDE6983D27255C93578050A57CEDFF1'


def search_for_words_in_excel():
    excel = pd.read_excel('C:\\Users\\82105\\Desktop\\신희승\\IIRTECH\\IIRTECH\\한국어 학습용 어휘 목록.xls', usecols=["단어", "등급"])
    words = []
    for value in excel.values:
        if value[1] == "A":  # 등급이 A로 나와있는 단어
            word = value[0]
            numbers = re.findall('\d', word)
            word = word.replace("".join(numbers), "")
            words.append(word)
    return list(set(words))


def search_for_words_in_opendict(words):
    # 마지막이 search면 사전 검색 오픈 API, view면 사전 내용 오픈 API입니다.
    defaultURL = 'https://opendict.korean.go.kr/api/view?'

    # key에 우리말샘 오픈 API 이용신청 후 받은 인증키를 넣어줍니다.
    key = 'key=' + opendict_cert_key

    # json 형식으로 반환
    req = "&req_type=json"
    i = 2
    for word in words:
        # q에 검색할 문자열을 넣어줍니다.
        index = 1
        while 1:
            if index >= 10:
                query = word + "0" + str(index)
            else:
                query = word + "00" + str(index)
            q = '&q=' + urllib.parse.quote(query)

            # 정렬 방식(기본값 dict)을 선택합니다. dict: 우리말샘순, popular: 많이 찾은 순, date: 새로 올린 순
            sort = '&sort=dict'
            full_url = defaultURL + key + req + q + sort

            request = urllib.request.Request(full_url)
            response = urllib.request.urlopen(request)
            rescode = response.getcode()

            if rescode == 200:
                text = response.read().decode('utf-8')
                if len(text) == 1:  # 페이지 없는 경우 예외처리
                    break
                json_object = json.loads(text)
                if 'item' not in json_object['channel']:
                    print("***********************************")
                    print(query)
                    print("***********************************")
                    break
                json_item = json_object['channel']['item']
                print(query)
                json_word = json_item['wordInfo']
                ws[columns_a[0] + str(i)] = json_word['word']  # 표제어
                ws[columns_a[3] + str(i)] = json_word['word_unit']  # 구성 단위
                ws[columns_a[4] + str(i)] = json_word['word_type']  # 고유어 여부
                if 'pronunciation_info' in json_word:
                    ws[columns_a[7] + str(i)] = str(json_word['pronunciation_info'])  # 발음

                json_sense = json_item['senseInfo']
                ws[columns_a[2] + str(i)] = json_sense['definition']  # 뜻풀이

                if 'pos' in json_sense:
                    ws[columns_a[14] + str(i)] = json_sense['pos']  # 품사

                ws[columns_a[15] + str(i)] = json_sense['type']  # 범주
                ws[columns_a[1] + str(i)] = json_sense['sense_no']  # 의미번호

                if 'original_language_info' in json_word:
                    original_language = ''
                    language_type = ''
                    for j in range(len(json_word['original_language_info'])):
                        json_original = json_word['original_language_info'][j]
                        original_language = original_language + json_original['original_language'] + ', '  # 원어
                        language_type = language_type + json_original['language_type'] + ', '
                    ws[columns_a[5] + str(i)] = original_language[0:len(original_language) - 2]  # 원어
                    ws[columns_a[6] + str(i)] = language_type[0:len(language_type) - 2]  # 언어

                if 'conju_info' in json_word:
                    conjugation = ''
                    conjugation_pro = ''
                    abbreviation = ''
                    abbreviation_pro = ''
                    for j in range(len(json_word['conju_info'])):
                        json_conju = json_word['conju_info'][j]
                        if 'conjugation_info' in json_conju:
                            json_conjugation = json_conju['conjugation_info']
                            conjugation = conjugation + json_conjugation['conjugation'] + ", "
                            if 'pronunciation_info' in json_conjugation:
                                conjugation_pro = conjugation_pro + str(json_conjugation['pronunciation_info']) + ", "
                        if 'abbreviation_info' in json_conju:
                            json_abbre = json_conju['abbreviation_info']
                            abbreviation = abbreviation + json_abbre['abbreviation'] + ', '
                            if 'pronunciation_info' in json_abbre:
                                abbreviation_pro = abbreviation_pro + str(json_abbre['pronunciation_info']) + ', '
                    ws[columns_a[8] + str(i)] = conjugation[0:len(conjugation) - 2]  # 활용
                    ws[columns_a[9] + str(i)] = conjugation_pro[0:len(conjugation_pro) - 2]  # 활용의 발음
                    ws[columns_a[10] + str(i)] = abbreviation[0:len(abbreviation) - 2]  # 준말
                    ws[columns_a[11] + str(i)] = abbreviation_pro[0:len(abbreviation_pro) - 2]  # 준말의 발음

                if 'origin' in json_word:
                    ws[columns_a[12] + str(i)] = json_word['origin']  # 어원

                if 'allomorph' in json_word:
                    ws[columns_a[13] + str(i)] = json_word['allomorph']  # 이형태

                if 'cat_info' in json_sense:
                    ws[columns_a[17] + str(i)] = str(json_sense['cat_info'])  # 전문분야

                if 'grammar_info' in json_sense:
                    ws[columns_a[16] + str(i)] = str(json_sense['grammar_info'])  # 문법

                if 'example_info' in json_sense:
                    example = ''
                    for j in range(len(json_sense['example_info'])):
                        json_exam = json_sense['example_info'][j]
                        example = example + json_exam['example'] + ', '
                    ws[columns_a[18] + str(i)] = example[0:len(example) - 2]  # 용례

                if 'relation_info' in json_sense:
                    relation = ''
                    relation_type = ''
                    for j in range(len(json_sense['relation_info'])):
                        relation = relation + json_sense['relation_info'][j]['word'] + ', '
                        relation_type = relation_type + json_sense['relation_info'][j]['type'] + ', '
                    ws[columns_a[19] + str(i)] = relation[0:len(relation) - 2]  # 관련 어휘
                    ws[columns_a[20] + str(i)] = relation_type[0:len(relation_type) - 2]  # 관련 어휘 유형

                if 'translation_info' in json_sense:
                    translation = ''
                    translation_type = ''
                    for j in range(len(json_sense['translation_info'])):
                        translation = translation + json_sense['translation_info'][j]['translation'] + ', '
                        translation_type = translation_type + json_sense['translation_info'][j]['language_type'] + ', '
                    ws[columns_a[21] + str(i)] = translation[0:len(translation) - 2]  # 대역어
                    ws[columns_a[22] + str(i)] = translation_type[0:len(translation_type) - 2]  # 대역어 언어

                if 'region_info' in json_sense:  # *****************************
                    ws[columns_a[23] + str(i)] = str(json_sense['region_info'])  # 방언지역

                if 'proverb_info' in json_word:
                    proverb = ''
                    proverb_defi = ''
                    proverb_type = ''
                    for j in range(len(json_word['proverb_info'])):
                        proverb = proverb + json_word['proverb_info'][j]['word']
                        proverb_defi = proverb_defi + json_word['proverb_info'][j]['definition']
                        proverb_type = proverb_type + json_word['proverb_info'][j]['type']
                    ws[columns_a[24] + str(i)] = proverb[0:len(proverb) - 2]              # 속담, 관용구
                    ws[columns_a[25] + str(i)] = proverb_defi[0:len(proverb_defi) - 2]  # 뜻풀이
                    ws[columns_a[26] + str(i)] = proverb_type[0:len(proverb_type) - 2]  # 유형

                ws[columns_a[27] + str(i)] = json_item['target_code']
                i += 1
                index += 1
            else:
                return print('response code:' + rescode)


wb = openpyxl.Workbook()

ws = wb.create_sheet('data')

i = 0
columns = ['표제어', '의미번호', '뜻', '구성단위', '고유어 여부', '원어', '언어', '발음', '활용', '활용의 발음', '준말', '준말의 발음', '어원', '이형태', '품사',
           '범주', '문법', '전문분야', '용례', '관련어휘', '관련어휘 유형', '대역어', '대역어 언어', '방언지역','속담/관용구', '뜻풀이', '유형', 'pk']

columns_a = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'M', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
             'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']

for column in columns:
    ws[columns_a[i] + '1'] = column
    i += 1

words = search_for_words_in_excel()
search_for_words_in_opendict(words)

wb.save("C:\\Users\\82105\\Desktop\\신희승\\IIRTECH\\IIRTECH\\data.xlsx")
