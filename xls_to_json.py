import openpyxl
import json
from collections import OrderedDict
import pyexcel as p

# excel_file = 'C:\Users\82105\Desktop\신희승\IIRTECH\IIRTECH\data.xlsx'
json_file = 'C:\\Users\\82105\\Desktop\\신희승\\IIRTECH\\IIRTECH\\data.json'
excel_file = 'C:\\Users\\82105\\Desktop\\신희승\\IIRTECH\\IIRTECH\\data.xlsx'

# xls to xlsx
# p.save_book_as(file_name = excel_file, dest_file_name= excel_file+"x")

wb = openpyxl.load_workbook(excel_file)

sheet = wb.worksheets[1]

# key_list 생성
col_max = sheet.max_column
row_max = sheet.max_row
key_list = ['word', 'sense_no', 'definition', 'word_unit', 'word_type', 'original_language', 'language_type',
            'pronunciation', 'conjugation', 'conjugation_pro', 'abbreviation', 'abbreviation_pro', 'origin',
            'allomorph', 'pos',
            'type', 'grammar', 'cat', 'example', 'source', 'origin', 'relation_word', 'translation', 'region', 'pk']
for i in range(1, col_max + 1):
    key_list.append(sheet.cell(row=1, column=i).value)

data_list = []
for row_num in range(2, row_max):
    data = OrderedDict()
    if row_num % 10 == 0:
        print(row_num)
    for i in range(1, col_max + 1):
        data[key_list[i - 1]] = sheet.cell(row=row_num, column=i).value
    data_list.append(data)

wb.close()

with open(json_file, 'w', encoding='utf-8') as fp:
    json.dump(data_list, fp, indent=4, ensure_ascii=False)
